"""
Catalogue Service — pivot table of available samples by participant.

The catalogue answers: "For each participant, how many aliquots of each
sample type do we have available?"

Output structure (pivot):
  Rows    = unique participants (PID + demographics)
  Columns = unique sample types (Serum, PBMC, ED Plasma, ...)
  Cells   = count of available aliquots for that participant × type

This is exactly what a researcher needs to decide which samples to request.

Key concept — building a pivot in pure Python:
  SQLAlchemy gives us flat rows. We reshape them into a dict of dicts:
    pivot[pid][sample_type] = count
  Then we convert that to a list of rows for the table/Excel writer.
  No pandas required — plain Python dicts are sufficient for this scale.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from sqlalchemy.orm import Session

from app.core.models.models import (
    Participant, Sample, SampleAliquot, Study
)


@dataclass
class CatalogueRow:
    """One row in the catalogue — one participant."""
    pid:        str
    study_code: str
    age:        Optional[int]
    gender:     Optional[str]
    disease:    Optional[str]
    cohort_name:     Optional[str]
    site_name:  Optional[str]
    # sample_counts is populated after pivot: {sample_type: count}
    sample_counts: Dict[str, int] = field(default_factory=dict)

    @property
    def total_aliquots(self) -> int:
        return sum(self.sample_counts.values())


class CatalogueService:

    def __init__(self, session: Session) -> None:
        self.session = session

    def generate(
        self,
        study_id: Optional[int] = None,
        sample_types: Optional[List[str]] = None,
        available_only: bool = True,
    ) -> Tuple[List[CatalogueRow], List[str]]:
        """
        Generate the sample catalogue.

        Returns:
          (rows, column_headers)
          rows           = list of CatalogueRow, one per participant
          column_headers = sorted list of sample type names (pivot columns)
        """
        # ── Query: all aliquots with their participant and sample info ─────
        q = (
            self.session.query(
                Participant, Sample, SampleAliquot
            )
            .join(Sample,       Sample.participant_id == Participant.id)
            .join(SampleAliquot, SampleAliquot.sample_id == Sample.id)
        )

        if study_id:
            q = q.filter(Participant.study_id == study_id)
        if available_only:
            q = q.filter(
                SampleAliquot.is_available == True,
                SampleAliquot.is_shipped   == False,
            )
        if sample_types:
            q = q.filter(Sample.sample_type.in_(sample_types))

        rows_raw = q.all()

        # ── Build pivot dict ──────────────────────────────────────────────
        # pivot[pid] = {sample_type: count}
        pivot: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))

        # participant_meta[pid] = CatalogueRow (without counts yet)
        participant_meta: Dict[str, CatalogueRow] = {}

        # study code lookup
        study_map: Dict[int, str] = {}

        for participant, sample, aliquot in rows_raw:
            pid = participant.pid

            # Cache study code
            if participant.study_id not in study_map:
                study = self.session.get(Study, participant.study_id)
                study_map[participant.study_id] = (
                    study.project_id_short if study else "?"
                )

            # Cache participant metadata
            if pid not in participant_meta:
                participant_meta[pid] = CatalogueRow(
                    pid=pid,
                    study_code=study_map[participant.study_id],
                    age=participant.age,
                    gender=participant.gender,
                    disease=participant.disease,
                    cohort_name=participant.cohort_name,
                    site_name=participant.site_name,
                )

            pivot[pid][sample.sample_type] += 1

        # ── Collect all unique sample types (pivot columns) ───────────────
        all_types: List[str] = sorted({
            sample_type
            for counts in pivot.values()
            for sample_type in counts
        })

        # ── Assemble final rows ───────────────────────────────────────────
        catalogue_rows: List[CatalogueRow] = []
        for pid, meta in sorted(participant_meta.items()):
            meta.sample_counts = {
                t: pivot[pid].get(t, 0) for t in all_types
            }
            catalogue_rows.append(meta)

        return catalogue_rows, all_types

    def export_to_excel(
        self,
        rows: List[CatalogueRow],
        sample_type_columns: List[str],
        filepath: str,
    ) -> None:
        """
        Write the catalogue to an Excel file with:
          - Frozen header row
          - Colour-coded availability (green = has samples, grey = none)
          - Auto-fitted column widths
          - Summary totals row at the bottom
        """
        import openpyxl
        from openpyxl.styles import (
            Alignment, Font, PatternFill, Border, Side
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sample Catalogue"

        # ── Styles ─────────────────────────────────────────────────────────
        hdr_fill   = PatternFill("solid", fgColor="1F4E79")
        hdr_font   = Font(color="FFFFFF", bold=True, size=10)
        grp_fill   = PatternFill("solid", fgColor="2E75B6")   # sample type group header
        grp_font   = Font(color="FFFFFF", bold=True, size=10)
        avail_fill = PatternFill("solid", fgColor="E2EFDA")   # green tint — has samples
        zero_fill  = PatternFill("solid", fgColor="F2F2F2")   # grey — no samples
        total_fill = PatternFill("solid", fgColor="FCE4D6")   # orange tint — totals row
        total_font = Font(bold=True, size=10)
        center     = Alignment(horizontal="center", vertical="center")
        thin       = Side(style="thin", color="CCCCCC")
        border     = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Header row ────────────────────────────────────────────────────
        demo_headers = ["PID", "Study", "Age", "gender", "Disease", "cohort_name", "Site", "Total"]
        all_headers  = demo_headers + sample_type_columns

        for col_idx, header in enumerate(all_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill  = hdr_fill if col_idx <= len(demo_headers) else grp_fill
            cell.font  = hdr_font if col_idx <= len(demo_headers) else grp_font
            cell.alignment = center
            cell.border = border

        # ── Data rows ─────────────────────────────────────────────────────
        for row_idx, row in enumerate(rows, start=2):
            demo_values = [
                row.pid, row.study_code, row.age, row.gender,
                row.disease, row.cohort_name, row.site_name, row.total_aliquots,
            ]
            # Demographic columns
            for col_idx, val in enumerate(demo_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = center
                cell.border = border

            # Sample type pivot columns
            for col_idx, stype in enumerate(sample_type_columns,
                                             start=len(demo_headers) + 1):
                count = row.sample_counts.get(stype, 0)
                cell  = ws.cell(row=row_idx, column=col_idx, value=count or None)
                cell.fill      = avail_fill if count > 0 else zero_fill
                cell.alignment = center
                cell.border    = border

        # ── Totals row ────────────────────────────────────────────────────
        total_row = len(rows) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
        for col_idx in range(1, len(all_headers) + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.fill   = total_fill
            cell.border = border
            cell.alignment = center

        # Sum each sample type column
        for col_idx, stype in enumerate(sample_type_columns,
                                         start=len(demo_headers) + 1):
            total = sum(r.sample_counts.get(stype, 0) for r in rows)
            ws.cell(row=total_row, column=col_idx, value=total or None).font = total_font

        # Grand total
        ws.cell(row=total_row, column=len(demo_headers),
                value=sum(r.total_aliquots for r in rows)).font = total_font

        # ── Column widths + freeze ─────────────────────────────────────────
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=6)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 28)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 24

        wb.save(filepath)
