"""
Excel bulk import service.

Handles parsing, validating, and importing participant/sample/aliquot data
from Excel files with structure:
  PID | Age | Gender | Population | Disease | Visit Code | Visit Time |
  Date Collected | Site Name | Visit Name | Sample Type | Cohort Name |
  Aliquot ID | Freezer / Tank | Container | Slot Position | Shelf | Rack |
  Position | Discrepancy Remark | Discrepancy For
"""

from __future__ import annotations

from datetime import datetime
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
import re
from typing import Callable, Optional
from dateutil.parser import parse, ParserError
import openpyxl

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.config import Gender, Population, Disease, Site, VisitName, SampleType, CohortName
from app.core.models.models import (
    Participant, Sample, SampleAliquot, Study, Freezer, Compartment,
    StorageRack, StorageDrawer, StorageBox, BoxPosition, AliquotLocation
)
from app.core.services.id_generator import generate_aliquot_id
from app.core.repositories.storage_repository import FreezerRepository

# Valid hierarchy values for upright freezers (Freezer 1 & 2)
VALID_SHELVES  = ["I", "II", "III", "IV"]
VALID_RACKS    = ["A", "B", "C", "D", "E", "F"]
VALID_DRAWERS  = ["01", "02", "03", "04", "05"]

# Cylindrical freezers (Freezer 3 & 4): racks 01-13, no shelf or drawer
VALID_CYLINDRICAL_RACKS           = [f"{i:02d}" for i in range(1, 14)]
CYLINDRICAL_SENTINEL_COMPARTMENT  = "CYLINDRICAL"
CYLINDRICAL_SENTINEL_DRAWER       = "01"

# Rows processed per flush+commit cycle during bulk import
IMPORT_BATCH_SIZE = 500

# Import-time aliases: maps lowercase raw value → canonical enum value.
_SITE_ALIASES: dict[str, str] = {
    "nari":      "ICMR-NARI",
    "icmr-nari": "ICMR-NARI",
    "nirt":      "NIRT",
    "icmr-nirt": "ICMR-NIRT",
}
_COHORT_ALIASES: dict[str, str] = {
    "hiv infected-adults": "HIV INFECTED-ADULT",
}
_DISEASE_ALIASES: dict[str, str] = {
    "infected w/o co-morbidity":      "Infected without co-morbidity",
    "infected without co-morbidity":  "Infected without co-morbidity",
    "unknown-screen failure":         "Unknown-Screen failure",
}
_SAMPLE_TYPE_ALIASES: dict[str, str] = {
    "hep plasma": "HEP Plasma",
}


def _apply_alias(raw: str, alias_map: dict[str, str]) -> str:
    return alias_map.get(raw.lower().strip(), raw)


def _clean_dot(val) -> Optional[str]:
    """Treat '.' placeholder cells as empty."""
    if val is None:
        return None
    s = str(val).strip()
    return None if s in ('', '.') else s


@dataclass
class ImportRow:
    """Parsed Excel row (0-indexed in file, but row_num is 1-indexed for display)."""
    row_num: int  # Display row number (2, 3, 4, ...)
    pid: Optional[str]
    age: Optional[int]
    gender: Optional[str]
    population: Optional[str]
    disease: Optional[str]
    visit_code: Optional[str]
    visit_time: Optional[str]
    date_collected: Optional[str]
    site_name: Optional[str]
    visit_name: Optional[str]
    sample_type: Optional[str]
    cohort_name: Optional[str]
    aliquot_id: Optional[str]
    freezer_name: Optional[str]
    container_name: Optional[str]
    slot_position: Optional[int]  # Sequential position 1-100 in box grid
    shelf_name: Optional[str]  # I, II, III, IV
    rack_drawer_combined: Optional[str]  # Format: "D-02" (Rack-Drawer)
    position: Optional[str]  # Will be converted from slot_position to letter+number (e.g., A1)
    discrepancy_remark: Optional[str]
    discrepancy_for: Optional[str]
    errors: list[str] = None

    def __post_init__(self):
        if self.errors is None:
            self.errors = []


class ExcelImportService:
    """Service for bulk importing participant/sample/aliquot data from Excel."""

    EXPECTED_HEADERS = [
        'PID', 'Age', 'Gender', 'Population', 'Disease', 'Visit Code', 'Visit Time',
        'Date Collected', 'Site Name', 'Visit Name', 'Sample Type', 'Cohort Name',
        'Aliquot ID', 'Freezer / Tank', 'Container', 'Slot Position', 'Shelf', 'Rack',
        'Position', 'Discrepancy Remark', 'Discrepancy For'
    ]

    def __init__(self, session: Session):
        self.session = session

    def load_and_validate_excel(self, filepath: str) -> tuple[list[ImportRow], list[str]]:
        """
        Load Excel file, parse rows, and validate each row.

        Uses read_only=True for streaming to avoid loading the full file into RAM.

        Returns:
          (validated_rows, header_errors)

        If header_errors is non-empty, validated_rows will be empty.
        Each row in validated_rows has an 'errors' list; if non-empty, row failed validation.
        """
        filepath = Path(filepath)
        if not filepath.exists():
            return [], [f"File not found: {filepath}"]

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            ws = wb.active
        except Exception as e:
            return [], [f"Failed to open Excel file: {e}"]

        try:
            # Validate headers by reading only the first row
            first_row_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            first_row = next(first_row_iter, ())
            actual_headers = [first_row[i] if i < len(first_row) else None for i in range(21)]

            mismatches = []
            for i, (exp, got) in enumerate(zip(self.EXPECTED_HEADERS, actual_headers)):
                e_norm = str(exp or '').lower().strip()
                g_norm = str(got or '').lower().strip()
                if e_norm != g_norm:
                    ratio = SequenceMatcher(None, e_norm, g_norm).ratio()
                    if ratio < 0.80:
                        mismatches.append(f"Column {i+1}: expected '{exp}', got '{got}'")
            if mismatches:
                return [], [
                    "Excel header mismatch:\n" + "\n".join(mismatches)
                ]

            # Stream rows from row 2 onward
            rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Ensure row has at least 21 columns
                row_data = list(row[:21]) + [None] * (21 - len(row[:21]))

                # Skip completely empty rows
                if all(v is None or str(v).strip() == '' for v in row_data):
                    continue

                import_row = ImportRow(
                    row_num=row_idx,
                    pid=row_data[0],
                    age=row_data[1],
                    gender=row_data[2],
                    population=row_data[3],
                    disease=row_data[4],
                    visit_code=row_data[5],
                    visit_time=row_data[6],
                    date_collected=row_data[7],
                    site_name=row_data[8],
                    visit_name=row_data[9],
                    sample_type=row_data[10],
                    cohort_name=row_data[11],
                    aliquot_id=row_data[12],
                    freezer_name=_clean_dot(row_data[13]),
                    container_name=_clean_dot(row_data[14]),
                    slot_position=row_data[15],
                    shelf_name=_clean_dot(row_data[16]),
                    rack_drawer_combined=_clean_dot(row_data[17]),
                    position=_clean_dot(row_data[18]),
                    discrepancy_remark=_clean_dot(row_data[19]),
                    discrepancy_for=_clean_dot(row_data[20]),
                )
                rows.append(import_row)

        finally:
            wb.close()

        # Validate each row
        for row in rows:
            self._validate_row(row)

        return rows, []

    def _validate_row(self, row: ImportRow) -> None:
        """Validate a single row and populate row.errors."""
        errors = []

        # PID (required)
        if not row.pid:
            errors.append("PID is required")
        elif not isinstance(row.pid, str):
            try:
                row.pid = str(row.pid).strip()
            except:
                errors.append("PID must be a string")

        # Age (optional, must be int if provided)
        if row.age is not None:
            try:
                row.age = int(row.age)
            except (ValueError, TypeError):
                errors.append(f"Age must be an integer, got '{row.age}'")

        # Gender
        if row.gender:
            raw = str(row.gender).strip()
            m = Gender(raw)
            if m is None:
                errors.append(f"Invalid Gender '{raw}'. Allowed: {[e.value for e in Gender]}")
            else:
                row.gender = m.value

        # Population
        if row.population:
            raw = str(row.population).strip()
            m = Population(raw)
            if m is None:
                errors.append(f"Invalid Population '{raw}'. Allowed: {[e.value for e in Population]}")
            else:
                row.population = m.value

        # Disease (alias first, then enum)
        if row.disease:
            raw = _apply_alias(str(row.disease).strip(), _DISEASE_ALIASES)
            m = Disease(raw)
            if m is None:
                errors.append(f"Invalid Disease '{row.disease}'. Allowed: {[e.value for e in Disease]}")
            else:
                row.disease = m.value

        # Site Name (alias first, then enum)
        if row.site_name:
            raw = _apply_alias(str(row.site_name).strip(), _SITE_ALIASES)
            m = Site(raw)
            if m is None:
                errors.append(f"Invalid Site '{row.site_name}'. Allowed: {[e.value for e in Site]}")
            else:
                row.site_name = m.value

        # Visit Name
        if row.visit_name:
            raw = str(row.visit_name).strip()
            m = VisitName(raw)
            if m is None:
                errors.append(f"Invalid Visit Name '{raw}'. Allowed: {[e.value for e in VisitName]}")
            else:
                row.visit_name = m.value

        # Sample Type (alias first, then enum)
        if row.sample_type:
            raw = _apply_alias(str(row.sample_type).strip(), _SAMPLE_TYPE_ALIASES)
            m = SampleType(raw)
            if m is None:
                errors.append(f"Invalid Sample Type '{row.sample_type}'. Allowed: {[e.value for e in SampleType]}")
            else:
                row.sample_type = m.value

        # Cohort Name (alias first, then enum)
        if row.cohort_name:
            raw = _apply_alias(str(row.cohort_name).strip(), _COHORT_ALIASES)
            m = CohortName(raw)
            if m is None:
                errors.append(f"Invalid Cohort Name '{row.cohort_name}'. Allowed: {[e.value for e in CohortName]}")
            else:
                row.cohort_name = m.value

        # Visit Code (optional, string)
        if row.visit_code:
            row.visit_code = str(row.visit_code).strip()

        # Visit Time (optional, HH:MM format)
        if row.visit_time:
            row.visit_time = str(row.visit_time).strip()
            if not self._is_valid_time(row.visit_time):
                errors.append(f"Visit Time must be HH:MM format, got '{row.visit_time}'")

        # Date Collected (optional, YYYY-MM-DD format)
        if row.date_collected:
            row.date_collected = str(row.date_collected).strip()
            if not self._is_valid_date(row.date_collected):
                errors.append(f"Date Collected must be YYYY-MM-DD format, got '{row.date_collected}'")

        # Storage hierarchy validation
        has_storage = any([row.freezer_name, row.container_name, row.shelf_name,
                           row.rack_drawer_combined, row.slot_position])

        if has_storage:
            missing = []
            if not row.freezer_name:        missing.append("Freezer / Tank")
            if not row.container_name:      missing.append("Container")
            if not row.rack_drawer_combined: missing.append("Rack")
            if not row.slot_position:       missing.append("Slot Position")
            if missing:
                errors.append(f"Storage incomplete. Missing: {', '.join(missing)}")
            else:
                is_cylindrical = not row.shelf_name

                if is_cylindrical:
                    rack_str = str(row.rack_drawer_combined).strip()
                    try:
                        rack_str = f"{int(float(rack_str)):02d}"
                        if rack_str not in VALID_CYLINDRICAL_RACKS:
                            errors.append(
                                f"Cylindrical rack must be 01–13, got '{rack_str}'"
                            )
                        else:
                            row.rack_drawer_combined = rack_str
                    except (ValueError, TypeError):
                        errors.append(
                            f"Cylindrical rack must be a number 1–13, got '{rack_str}'"
                        )
                else:
                    row.shelf_name = str(row.shelf_name).strip()
                    if row.shelf_name not in VALID_SHELVES:
                        errors.append(
                            f"Invalid Shelf '{row.shelf_name}'. Allowed: {VALID_SHELVES}"
                        )

                    rack_drawer = str(row.rack_drawer_combined).strip()
                    parts = rack_drawer.split('-')
                    if len(parts) != 2:
                        errors.append(
                            f"Rack format invalid '{rack_drawer}'. Expected: 'A-01'"
                        )
                    else:
                        rack_name   = parts[0].upper()
                        drawer_name = parts[1].strip()
                        if rack_name not in VALID_RACKS:
                            errors.append(
                                f"Invalid Rack '{rack_name}'. Allowed: {VALID_RACKS}"
                            )
                        if drawer_name not in VALID_DRAWERS:
                            errors.append(
                                f"Invalid Drawer '{drawer_name}'. Allowed: {VALID_DRAWERS}"
                            )
                        row.rack_drawer_combined = f"{rack_name}-{drawer_name}"

                try:
                    slot_pos = int(row.slot_position)
                    if slot_pos < 1 or slot_pos > 100:
                        errors.append(
                            f"Slot Position must be 1–100, got '{slot_pos}'"
                        )
                    else:
                        row.position = self._convert_position_number_to_format(slot_pos)
                except (ValueError, TypeError):
                    errors.append(
                        f"Slot Position must be a number, got '{row.slot_position}'"
                    )

        row.errors = errors

    def _is_valid_time(self, time_str: str) -> bool:
        """Check if time_str matches visit time codes: SCR, SCR (NA), M0–M36."""
        pattern = re.compile(r'^(SCR|SCR \(NA\)|M([0-9]|[12][0-9]|3[0-6]))$')
        try:
            if pattern.match(time_str):
                return True
        except ValueError:
            return False

    def _is_valid_date(self, date_str: str) -> bool:
        """Check if date_str is valid YYYY-MM-DD format."""
        if not isinstance(date_str, str):
                date_str = str(date_str).strip()
        try:
            parse(date_str, dayfirst=True)
            return True
        except (ParserError, ValueError, TypeError):
            return False

    def _is_valid_position_format(self, position: str) -> bool:
        """Check if position format is valid (letter A-J + number 1-10, e.g., A1, J10)."""
        try:
            if not position or len(position) < 2:
                return False
            col_letter = position[0].upper()
            row_num = int(position[1:])
            if not ('A' <= col_letter <= 'J') or row_num < 1 or row_num > 10:
                return False
            return True
        except (ValueError, TypeError):
            return False

    def _convert_position_number_to_format(self, position_number: int) -> str:
        """
        Convert sequential position number (1-100) to grid format (A1-J10).
        Position 1-10 = A1-J1, 11-20 = A2-J2, ..., 91-100 = A10-J10
        """
        zero_indexed = position_number - 1
        row = (zero_indexed // 10) + 1  # 1-10
        col_idx = zero_indexed % 10     # 0-9
        col_letter = chr(65 + col_idx)  # A-J
        return f"{col_letter}{row}"

    def _position_to_row_col(self, position: str) -> tuple[Optional[int], Optional[int]]:
        """
        Convert position string (e.g., "A1", "B5", "J10") to (row, col) indices.

        Returns: (0-based row, 0-based col) or (None, None) if invalid
        """
        if not position or len(position) < 2:
            return None, None

        try:
            col_letter = position[0].upper()
            row_num = int(position[1:])
            col = ord(col_letter) - ord('A')
            row = row_num - 1
            if row < 0 or row >= 10 or col < 0 or col >= 10:
                return None, None
            return row, col
        except (ValueError, AttributeError):
            return None, None

    # ── Storage hierarchy helpers (original, used by non-bulk paths) ─────────

    def _get_or_create_storage_hierarchy(
        self, freezer_name: str, container_name: str,
        shelf_name: str, rack_drawer_combined: str
    ) -> tuple[Freezer, Compartment, StorageRack, StorageDrawer, StorageBox]:
        """Upright freezer hierarchy (Freezer 1 & 2)."""
        freezer_repo = FreezerRepository(self.session)
        rack_letter, drawer_number = rack_drawer_combined.split('-')

        freezer = freezer_repo.get_by_name(freezer_name)
        if not freezer:
            freezer = Freezer(name=freezer_name)
            self.session.add(freezer)
            self.session.flush()

        compartment = self.session.query(Compartment).filter(
            Compartment.name == shelf_name,
            Compartment.freezer_id == freezer.id,
        ).first()
        if not compartment:
            compartment = Compartment(name=shelf_name, freezer_id=freezer.id)
            self.session.add(compartment)
            self.session.flush()
            for rack_val in VALID_RACKS:
                self.session.add(StorageRack(name=rack_val, compartment_id=compartment.id))
            self.session.flush()

        rack = self.session.query(StorageRack).filter(
            StorageRack.name == rack_letter,
            StorageRack.compartment_id == compartment.id,
        ).first()
        if not rack:
            rack = StorageRack(name=rack_letter, compartment_id=compartment.id)
            self.session.add(rack)
            self.session.flush()
            for drawer_val in VALID_DRAWERS:
                self.session.add(StorageDrawer(name=drawer_val, rack_id=rack.id))
            self.session.flush()

        drawer = self.session.query(StorageDrawer).filter(
            StorageDrawer.name == drawer_number,
            StorageDrawer.rack_id == rack.id,
        ).first()
        if not drawer:
            drawer = StorageDrawer(name=drawer_number, rack_id=rack.id)
            self.session.add(drawer)
            self.session.flush()

        box = self.session.query(StorageBox).filter(
            StorageBox.name == container_name,
            StorageBox.drawer_id == drawer.id,
        ).first()
        if not box:
            box = StorageBox(name=container_name, drawer_id=drawer.id, rows=10, cols=10)
            self.session.add(box)
            self.session.flush()
            for r in range(10):
                for c in range(10):
                    self.session.add(BoxPosition(box_id=box.id, row=r, col=c))
            self.session.flush()

        return freezer, compartment, rack, drawer, box

    def _get_or_create_storage_hierarchy_cylindrical(
        self, freezer_name: str, container_name: str, rack_number: str
    ) -> tuple[Freezer, Compartment, StorageRack, StorageDrawer, StorageBox]:
        """Cylindrical freezer hierarchy (Freezer 3 & 4)."""
        freezer_repo = FreezerRepository(self.session)

        freezer = freezer_repo.get_by_name(freezer_name)
        if not freezer:
            freezer = Freezer(name=freezer_name)
            self.session.add(freezer)
            self.session.flush()

        compartment = self.session.query(Compartment).filter(
            Compartment.name == CYLINDRICAL_SENTINEL_COMPARTMENT,
            Compartment.freezer_id == freezer.id,
        ).first()
        if not compartment:
            compartment = Compartment(
                name=CYLINDRICAL_SENTINEL_COMPARTMENT, freezer_id=freezer.id
            )
            self.session.add(compartment)
            self.session.flush()

        rack = self.session.query(StorageRack).filter(
            StorageRack.name == rack_number,
            StorageRack.compartment_id == compartment.id,
        ).first()
        if not rack:
            rack = StorageRack(name=rack_number, compartment_id=compartment.id)
            self.session.add(rack)
            self.session.flush()

        drawer = self.session.query(StorageDrawer).filter(
            StorageDrawer.name == CYLINDRICAL_SENTINEL_DRAWER,
            StorageDrawer.rack_id == rack.id,
        ).first()
        if not drawer:
            drawer = StorageDrawer(
                name=CYLINDRICAL_SENTINEL_DRAWER, rack_id=rack.id
            )
            self.session.add(drawer)
            self.session.flush()

        box = self.session.query(StorageBox).filter(
            StorageBox.name == container_name,
            StorageBox.drawer_id == drawer.id,
        ).first()
        if not box:
            box = StorageBox(name=container_name, drawer_id=drawer.id, rows=10, cols=10)
            self.session.add(box)
            self.session.flush()
            for r in range(10):
                for c in range(10):
                    self.session.add(BoxPosition(box_id=box.id, row=r, col=c))
            self.session.flush()

        return freezer, compartment, rack, drawer, box

    # ── Optimised bulk-import storage cache ──────────────────────────────────

    def _get_box_cached(
        self,
        row: ImportRow,
        freezer_cache: dict,
        compartment_cache: dict,
        rack_cache: dict,
        drawer_cache: dict,
        box_cache: dict,
        position_cache: dict,
    ) -> Optional[tuple[StorageBox, str, str, str]]:
        """
        Get or create the StorageBox for a row, using in-memory caches to avoid
        repeated DB queries for the same hierarchy nodes.

        Returns (box, compartment_name, rack_name, drawer_name) or None on error.
        """
        is_cylindrical = not row.shelf_name

        # ── Freezer ──────────────────────────────────────────────────────────
        freezer = freezer_cache.get(row.freezer_name)
        if not freezer:
            freezer = (
                self.session.query(Freezer)
                .filter(Freezer.name == row.freezer_name)
                .first()
            )
            if not freezer:
                freezer = Freezer(name=row.freezer_name)
                self.session.add(freezer)
                self.session.flush()
            freezer_cache[row.freezer_name] = freezer

        # ── Compartment (shelf for upright, sentinel for cylindrical) ─────────
        if is_cylindrical:
            compartment_key_name = CYLINDRICAL_SENTINEL_COMPARTMENT
            rack_key_name        = row.rack_drawer_combined
            drawer_key_name      = CYLINDRICAL_SENTINEL_DRAWER
        else:
            compartment_key_name = row.shelf_name
            parts = row.rack_drawer_combined.split('-')
            rack_key_name   = parts[0]
            drawer_key_name = parts[1]

        comp_key = (freezer.id, compartment_key_name)
        compartment = compartment_cache.get(comp_key)
        if not compartment:
            compartment = (
                self.session.query(Compartment)
                .filter(
                    Compartment.name == compartment_key_name,
                    Compartment.freezer_id == freezer.id,
                )
                .first()
            )
            if not compartment:
                compartment = Compartment(name=compartment_key_name, freezer_id=freezer.id)
                self.session.add(compartment)
                self.session.flush()
                if not is_cylindrical:
                    for rv in VALID_RACKS:
                        self.session.add(StorageRack(name=rv, compartment_id=compartment.id))
                    self.session.flush()
            compartment_cache[comp_key] = compartment

        # ── Rack ──────────────────────────────────────────────────────────────
        rack_key = (compartment.id, rack_key_name)
        rack = rack_cache.get(rack_key)
        if not rack:
            rack = (
                self.session.query(StorageRack)
                .filter(
                    StorageRack.name == rack_key_name,
                    StorageRack.compartment_id == compartment.id,
                )
                .first()
            )
            if not rack:
                rack = StorageRack(name=rack_key_name, compartment_id=compartment.id)
                self.session.add(rack)
                self.session.flush()
                if not is_cylindrical:
                    for dv in VALID_DRAWERS:
                        self.session.add(StorageDrawer(name=dv, rack_id=rack.id))
                    self.session.flush()
            rack_cache[rack_key] = rack

        # ── Drawer ────────────────────────────────────────────────────────────
        drawer_key = (rack.id, drawer_key_name)
        drawer = drawer_cache.get(drawer_key)
        if not drawer:
            drawer = (
                self.session.query(StorageDrawer)
                .filter(
                    StorageDrawer.name == drawer_key_name,
                    StorageDrawer.rack_id == rack.id,
                )
                .first()
            )
            if not drawer:
                drawer = StorageDrawer(name=drawer_key_name, rack_id=rack.id)
                self.session.add(drawer)
                self.session.flush()
            drawer_cache[drawer_key] = drawer

        # ── Box ───────────────────────────────────────────────────────────────
        box_key = (drawer.id, row.container_name)
        box = box_cache.get(box_key)
        if not box:
            box = (
                self.session.query(StorageBox)
                .filter(
                    StorageBox.name == row.container_name,
                    StorageBox.drawer_id == drawer.id,
                )
                .first()
            )
            if not box:
                box = StorageBox(name=row.container_name, drawer_id=drawer.id, rows=10, cols=10)
                self.session.add(box)
                self.session.flush()
                # Bulk-create all 100 positions in one shot
                self.session.bulk_save_objects([
                    BoxPosition(box_id=box.id, row=r, col=c)
                    for r in range(10) for c in range(10)
                ])
                self.session.flush()

            box_cache[box_key] = box

            # Populate position cache for this box (one query, 100 results)
            for pos in self.session.query(BoxPosition).filter(BoxPosition.box_id == box.id).all():
                position_cache[(box.id, pos.row, pos.col)] = pos.id

        return box, compartment_key_name, rack_key_name, drawer_key_name

    # ── Public import entry point ────────────────────────────────────────────

    def import_rows(
        self,
        rows: list[ImportRow],
        study_id: int,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> tuple[int, Optional[str]]:
        """
        Import validated rows into the database.

        Optimisations vs. the naïve approach:
          - Pre-computes the starting sample serial with a single MAX() query.
          - Caches existing participants in memory to avoid per-row SELECT.
          - Caches storage hierarchy nodes to avoid repeated DB round-trips.
          - Uses ORM relationships (not raw integer FKs) so SQLAlchemy can batch
            INSERT ordering within each flush call.
          - Flushes and commits once per IMPORT_BATCH_SIZE rows instead of 3×
            per row.

        Args:
            rows:              Pre-validated ImportRow list (no rows with errors).
            study_id:          Target study primary key.
            progress_callback: Optional (processed, total) → None callback.

        Returns:
            (count_created, error_message)
        """
        rows_with_errors = [r for r in rows if r.errors]
        if rows_with_errors:
            error_list = "\n".join([
                f"Row {r.row_num}: {'; '.join(r.errors)}"
                for r in rows_with_errors
            ])
            return 0, f"Validation failed:\n{error_list}"

        try:
            study = self.session.query(Study).filter(Study.id == study_id).one()

            # ── Pre-compute starting sample serial (one MAX query) ────────────
            import datetime as dt
            year_short = str(dt.date.today().year)[-2:]
            prefix = f"{study.project_id_short}-{year_short}-"

            max_id = (
                self.session.query(func.max(Sample.sample_id))
                .filter(Sample.sample_id.like(f"{prefix}%"))
                .scalar()
            )
            next_serial = 1
            if max_id:
                try:
                    next_serial = int(max_id.replace(prefix, "")) + 1
                except ValueError:
                    pass

            # ── Pre-load existing participants → {pid: participant_id} ────────
            existing_participant_ids: dict[str, int] = {}
            for p in (
                self.session.query(Participant.pid, Participant.id)
                .filter(Participant.study_id == study_id)
                .all()
            ):
                existing_participant_ids[p.pid] = p.id

            # ── Storage hierarchy in-memory caches ────────────────────────────
            freezer_cache: dict     = {}
            compartment_cache: dict = {}
            rack_cache: dict        = {}
            drawer_cache: dict      = {}
            box_cache: dict         = {}
            position_cache: dict    = {}  # (box_id, row, col) → position_id

            created_count = 0
            total = len(rows)

            for batch_start in range(0, total, IMPORT_BATCH_SIZE):
                batch = rows[batch_start: batch_start + IMPORT_BATCH_SIZE]

                # Track new participants created in this batch (pid → ORM object)
                # so we can capture their IDs after the batch flush.
                new_participants_in_batch: dict[str, Participant] = {}

                for row in batch:
                    # ── Participant ───────────────────────────────────────────
                    pid = str(row.pid).strip() if row.pid else row.pid
                    p_id = existing_participant_ids.get(pid)

                    if p_id is not None:
                        # Use integer FK — no ORM object needed
                        participant_ref = None
                        participant_id  = p_id
                    else:
                        # Check if we already staged a new participant in this batch
                        if pid in new_participants_in_batch:
                            participant_ref = new_participants_in_batch[pid]
                            participant_id  = None
                        else:
                            participant_ref = Participant(
                                pid=pid,
                                study_id=study_id,
                                age=row.age,
                                gender=row.gender,
                                population=row.population,
                                disease=row.disease,
                                site_name=row.site_name,
                                cohort_name=row.cohort_name,
                                notes=row.discrepancy_remark,
                            )
                            self.session.add(participant_ref)
                            new_participants_in_batch[pid] = participant_ref
                            participant_id = None

                    # ── Sample ────────────────────────────────────────────────
                    parsed_date = self._parse_date(row.date_collected)
                    sample_id_str = f"{prefix}{next_serial}"
                    next_serial += 1

                    if participant_id is not None:
                        sample = Sample(
                            sample_id=sample_id_str,
                            participant_id=participant_id,
                            study_id=study_id,
                            sample_type=row.sample_type,
                            visit_time=row.visit_time,
                            collection_date=parsed_date,
                            visit_code=row.visit_code,
                            visit_name=row.visit_name,
                        )
                    else:
                        # Link via ORM relationship — SQLAlchemy resolves FK on flush
                        sample = Sample(
                            sample_id=sample_id_str,
                            participant=participant_ref,
                            study_id=study_id,
                            sample_type=row.sample_type,
                            visit_time=row.visit_time,
                            collection_date=parsed_date,
                            visit_code=row.visit_code,
                            visit_name=row.visit_name,
                        )
                    self.session.add(sample)

                    # ── Aliquot ───────────────────────────────────────────────
                    aliquot_id_str = generate_aliquot_id(sample_id_str, 1)
                    aliquot = SampleAliquot(
                        aliquot_id=aliquot_id_str,
                        sample=sample,          # relationship — resolved on flush
                        aliquot_number=1,
                        discrepancy_remark=row.discrepancy_remark,
                        discrepancy_field=row.discrepancy_for,
                    )
                    self.session.add(aliquot)

                    # ── Storage location ──────────────────────────────────────
                    if row.freezer_name:
                        result = self._get_box_cached(
                            row,
                            freezer_cache, compartment_cache,
                            rack_cache, drawer_cache,
                            box_cache, position_cache,
                        )
                        if result and row.position:
                            box, compartment_name, rack_name, drawer_name = result
                            grid_row, grid_col = self._position_to_row_col(row.position)
                            if grid_row is not None:
                                pos_id = position_cache.get((box.id, grid_row, grid_col))
                                if pos_id:
                                    location = AliquotLocation(
                                        aliquot=aliquot,    # relationship — resolved on flush
                                        position_id=pos_id,
                                        freezer_name=row.freezer_name,
                                        compartment_name=compartment_name,
                                        rack_name=rack_name,
                                        drawer_name=drawer_name,
                                        box_name=row.container_name,
                                    )
                                    self.session.add(location)

                    created_count += 1

                # ── Batch flush: SQLAlchemy resolves FK ordering automatically ─
                self.session.flush()

                # Capture IDs of newly created participants before commit
                for pid, p_obj in new_participants_in_batch.items():
                    existing_participant_ids[pid] = p_obj.id

                self.session.commit()

                if progress_callback:
                    progress_callback(min(batch_start + IMPORT_BATCH_SIZE, total), total)

            return created_count, None

        except Exception as e:
            self.session.rollback()
            return 0, f"Import failed: {str(e)}"

    @staticmethod
    def _parse_date(date_collected) -> Optional[datetime]:
        """Parse date_collected field to datetime, or None."""
        if not date_collected:
            return None
        if isinstance(date_collected, str):
            try:
                return parse(date_collected)
            except (ParserError, ValueError, TypeError):
                return None
        if isinstance(date_collected, datetime):
            return date_collected
        if hasattr(date_collected, 'date'):
            return datetime.combine(date_collected, datetime.min.time())
        return None
