"""
Participant Service — registration, editing, and bulk Excel import.

Key concept — bulk Excel import strategy:
  1. Read the file with openpyxl (doesn't require Excel installed)
  2. Validate every row before touching the DB
  3. Collect all errors upfront (don't stop at first error)
  4. Only if zero errors → insert everything in one transaction
  This is called "validate-then-commit" and prevents partial imports.
"""

from __future__ import annotations

from typing import List, Optional, Tuple

from sqlalchemy.orm import Session

from app.core.models.models import Participant, Study
from app.core.repositories.participant_repository import ParticipantRepository
from app.core.repositories.study_repository import StudyRepository
from app.core.services.audit_service import log
from app.core.services.auth_service import app_session
from app.core.services.id_generator import generate_sample_id
from app.config import AuditAction


# ── Expected Excel columns for bulk import ─────────────────────────────────
#
# The lab uploads Excel files in a fixed template.
# We map column header → model field here so it's easy to update later.

EXCEL_COLUMN_MAP = {
    "initials":   "initials",
    "age":        "age",
    "sex":        "sex",
    "dob":        "date_of_birth",
    "cohort":     "cohort",
    "population": "population",
    "site":       "site_name",
    "notes":      "notes",
}

REQUIRED_EXCEL_COLUMNS = {"initials", "age", "sex"}


class ParticipantService:

    def __init__(self, session: Session) -> None:
        self.session = session
        self.repo = ParticipantRepository(session)
        self.study_repo = StudyRepository(session)

    # ── Helpers ────────────────────────────────────────────────────────────

    def _next_pid(self, study: Study) -> str:
        """
        PIDs follow the same format as sample IDs: COH-26-1, COH-26-2 ...
        We reuse the same counter scoped to (study × year).
        """
        import datetime as dt
        year_short = str(dt.date.today().year)[-2:]
        prefix = f"{study.project_id_short}-{year_short}-"
        existing = (
            self.session.query(Participant)
            .filter(Participant.pid.like(f"{prefix}%"))
            .all()
        )
        max_serial = 0
        for p in existing:
            try:
                serial = int(p.pid.replace(prefix, ""))
                max_serial = max(max_serial, serial)
            except ValueError:
                pass
        return f"{prefix}{max_serial + 1}"

    # ── Create ─────────────────────────────────────────────────────────────

    def get_by_id(self, participant_id: int) -> Optional[Participant]:
        """Get participant by ID."""
        return self.repo.get_by_id(participant_id)

    def register_participant(
        self,
        study_id: int,
        initials: str,
        age: Optional[int] = None,
        sex: Optional[str] = None,
        date_of_birth=None,
        cohort: str = "",
        population: str = "",
        disease: str = "",
        comorbidity: str = "",
        site_name: str = "",
        notes: str = "",
    ) -> Tuple[bool, str, Optional[Participant]]:
        """Register a single participant and auto-generate their PID."""
        app_session.require("participant.create")

        study = self.study_repo.get_by_id(study_id)
        if not study:
            return False, "Study not found.", None
        if study.is_locked:
            return False, "Study is locked.", None
        if not initials.strip():
            return False, "Initials are required.", None

        pid = self._next_pid(study)

        participant = Participant(
            pid=pid,
            study_id=study_id,
            initials=initials.strip().upper(),
            age=age,
            sex=sex,
            date_of_birth=date_of_birth,
            cohort=cohort,
            population=population,
            disease=disease,
            comorbidity=comorbidity,
            site_name=site_name,
            notes=notes,
        )
        self.repo.add(participant)
        log(self.session, AuditAction.CREATE, "Participant", pid,
            f"Participant '{pid}' registered in study '{study.project_id_short}'.")

        return True, f"Participant registered with PID: {pid}", participant

    # ── Update ─────────────────────────────────────────────────────────────

    def create_participant(
        self,
        pid: str,
        study_id: int,
        age: Optional[int] = None,
        gender: Optional[str] = None,
        population: Optional[str] = None,
        disease: Optional[str] = None,
        site_name: Optional[str] = None,
        cohort_name: Optional[str] = None,
        comorbidity: str = "",
        notes: str = "",
    ) -> Tuple[bool, str, Optional[Participant]]:
        """Create a participant with user-provided PID (not auto-generated)."""
        app_session.require("participant.create")

        pid = pid.strip()
        if not pid:
            return False, "PID is required.", None

        study = self.study_repo.get_by_id(study_id)
        if not study:
            return False, "Study not found.", None
        if study.is_locked:
            return False, "Study is locked.", None

        if self.repo.get_by_pid(pid):
            return False, f"PID '{pid}' already exists.", None

        participant = Participant(
            pid=pid,
            study_id=study_id,
            age=age,
            gender=gender,
            population=population,
            disease=disease,
            site_name=site_name,
            cohort_name=cohort_name,
            comorbidity=comorbidity.strip(),
            notes=notes.strip(),
        )
        self.repo.add(participant)
        log(self.session, AuditAction.CREATE, "Participant", pid,
            f"Participant '{pid}' created in study '{study.project_id_short}'.")

        return True, f"Participant created with PID: {pid}", participant

    def update_participant(
        self,
        participant_id: int,
        age: Optional[int] = None,
        gender: Optional[str] = None,
        population: Optional[str] = None,
        disease: Optional[str] = None,
        site_name: Optional[str] = None,
        cohort_name: Optional[str] = None,
        comorbidity: str = "",
        notes: str = "",
        reason: str = "Updated via UI",
    ) -> Tuple[bool, str]:
        app_session.require("participant.edit")

        participant = self.repo.get_by_id(participant_id)
        if not participant:
            return False, "Participant not found."

        # Update fields
        if age is not None:
            participant.age = age
        if gender is not None:
            participant.gender = gender
        if population is not None:
            participant.population = population
        if disease is not None:
            participant.disease = disease
        if site_name is not None:
            participant.site_name = site_name
        if cohort_name is not None:
            participant.cohort_name = cohort_name
        if comorbidity:
            participant.comorbidity = comorbidity.strip()
        if notes:
            participant.notes = notes.strip()

        participant.edit_reason = reason

        self.repo.update(participant)
        log(self.session, AuditAction.UPDATE, "Participant",
            participant.pid, f"Updated. Reason: {reason}")
        return True, "Participant updated."

    # ── Search ─────────────────────────────────────────────────────────────

    def search(self, filters: dict, use_or: bool = False,
               page: int = 1, page_size: int = 50):
        return self.repo.search(filters, use_or, page, page_size)

    # ── Bulk Excel import ──────────────────────────────────────────────────

    def import_from_excel(
        self,
        filepath: str,
        study_id: int,
        sheet_name: str = None,
    ) -> Tuple[bool, str, dict]:
        """
        Import participants from an Excel file.

        Returns (success, summary_message, details_dict)
        details_dict = {
            "imported": int,
            "skipped":  int,
            "errors":   [(row_number, error_message), ...]
        }

        Key concept — openpyxl:
          We use openpyxl in read-only mode (read_only=True) for performance.
          iter_rows(values_only=True) gives us plain Python values — no Cell objects.
          The first row is assumed to be headers.
        """
        import openpyxl

        app_session.require("participant.create")

        study = self.study_repo.get_by_id(study_id)
        if not study:
            return False, "Study not found.", {}

        # ── 1. Load workbook ───────────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
        except Exception as e:
            return False, f"Could not open file: {e}", {}

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return False, "File is empty.", {}

        # ── 2. Parse headers ───────────────────────────────────────────────
        # Normalise to lowercase so column order doesn't matter
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        missing = REQUIRED_EXCEL_COLUMNS - set(headers)
        if missing:
            return False, f"Missing required columns: {', '.join(missing)}", {}

        # ── 3. Validate all rows first ─────────────────────────────────────
        errors = []
        valid_rows = []

        for row_idx, row in enumerate(rows[1:], start=2):   # row 1 = headers
            row_data = dict(zip(headers, row))
            row_errors = []

            initials = str(row_data.get("initials", "") or "").strip()
            if not initials:
                row_errors.append("initials is empty")

            age_raw = row_data.get("age")
            age = None
            if age_raw is not None:
                try:
                    age = int(age_raw)
                    if not (0 < age < 150):
                        row_errors.append(f"age '{age}' out of range")
                except (ValueError, TypeError):
                    row_errors.append(f"age '{age_raw}' is not a number")

            sex = str(row_data.get("sex", "") or "").strip().capitalize()
            if sex and sex not in ("Male", "Female", "Other"):
                row_errors.append(f"sex '{sex}' must be Male/Female/Other")

            if row_errors:
                errors.append((row_idx, "; ".join(row_errors)))
            else:
                valid_rows.append({
                    "initials": initials,
                    "age": age,
                    "sex": sex or None,
                    "cohort": str(row_data.get("cohort", "") or ""),
                    "population": str(row_data.get("population", "") or ""),
                    "site_name": str(row_data.get("site", "") or ""),
                    "notes": str(row_data.get("notes", "") or ""),
                })

        # ── 4. Abort if any errors ─────────────────────────────────────────
        if errors:
            return False, f"{len(errors)} row(s) have errors. No data imported.", {
                "imported": 0, "skipped": 0, "errors": errors,
            }

        # ── 5. Insert all valid rows ───────────────────────────────────────
        imported = 0
        for row_data in valid_rows:
            ok, msg, _ = self.register_participant(study_id=study_id, **row_data)
            if ok:
                imported += 1

        wb.close()

        log(self.session, AuditAction.CREATE, "Participant", None,
            f"Bulk import: {imported} participants added to '{study.project_id_short}'.")

        return True, f"Successfully imported {imported} participants.", {
            "imported": imported, "skipped": 0, "errors": [],
        }