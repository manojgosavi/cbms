"""
Participant Tab — searchable list of participants with toolbar.
"""

from __future__ import annotations

from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (
    QAbstractItemView, QComboBox, QHBoxLayout, QHeaderView,
    QLabel, QLineEdit, QMessageBox, QPushButton,
    QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget,
)

from app.core.models.database import get_session
from app.core.services.study_service import StudyService
from app.core.services.participant_service import ParticipantService
from app.ui.widgets.pagination_bar import PaginationBar
from app.utils.exception_handler import slot_safe

PAGE_SIZE = 100


class ParticipantTab(QWidget):

    COLUMNS = ["PID", "Study", "Age", "Gender", "Population", "Disease",
               "Site", "Cohort Name", "Visit Code", "Notes", "Registered"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._studies = []
        self._page = 1
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)

        # ── Filter bar ─────────────────────────────────────────────────────
        filter_row = QHBoxLayout()

        self._study_filter = QComboBox()
        self._study_filter.addItem("All Studies", None)
        self._study_filter.currentIndexChanged.connect(self._on_filter_changed)

        self._pid_search = QLineEdit()
        self._pid_search.setPlaceholderText("Search by PID…")
        self._pid_search.setMaximumWidth(180)
        self._pid_search.textChanged.connect(self._on_filter_changed)

        filter_row.addWidget(QLabel("Study:"))
        filter_row.addWidget(self._study_filter)
        filter_row.addWidget(QLabel("  PID:"))
        filter_row.addWidget(self._pid_search)
        filter_row.addStretch()
        layout.addLayout(filter_row)

        # ── Toolbar ────────────────────────────────────────────────────────
        toolbar = QHBoxLayout()

        self._btn_new = QPushButton("＋  Register Participant")
        self._btn_new.clicked.connect(self._on_new)

        self._btn_import = QPushButton("📥  Import from Excel")
        self._btn_import.clicked.connect(self._on_import)

        self._btn_edit = QPushButton("✎  Edit")
        self._btn_edit.setEnabled(False)
        self._btn_edit.clicked.connect(self._on_edit)

        self._btn_export = QPushButton("📤  Export to Excel")
        self._btn_export.setEnabled(False)
        self._btn_export.clicked.connect(self._on_export)

        self._lbl_count = QLabel("0 participants")
        self._lbl_count.setStyleSheet("color: grey;")

        toolbar.addWidget(self._btn_new)
        toolbar.addWidget(self._btn_import)
        toolbar.addWidget(self._btn_edit)
        toolbar.addWidget(self._btn_export)
        toolbar.addStretch()
        toolbar.addWidget(self._lbl_count)
        layout.addLayout(toolbar)

        # ── Table ──────────────────────────────────────────────────────────
        self._table = QTableWidget()
        self._table.setColumnCount(len(self.COLUMNS))
        self._table.setHorizontalHeaderLabels(self.COLUMNS)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setAlternatingRowColors(True)

        header = self._table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.Stretch)  # Notes

        self._table.verticalHeader().setVisible(False)
        self._table.itemSelectionChanged.connect(
            lambda: self._btn_edit.setEnabled(bool(self._table.selectedItems()))
        )
        layout.addWidget(self._table)

        # ── Pagination bar ─────────────────────────────────────────────────
        self._pagination = PaginationBar()
        self._pagination.page_changed.connect(self._on_page_changed)
        layout.addWidget(self._pagination)

    def _on_filter_changed(self):
        self._page = 1
        self.refresh()

    def _on_page_changed(self, page: int):
        self._page = page
        self.refresh()

    def _load_studies(self):
        with get_session() as session:
            service = StudyService(session)
            self._studies = [(s.id, s.project_id_short) for s in service.get_all_active()]

        current = self._study_filter.currentData()
        self._study_filter.blockSignals(True)
        self._study_filter.clear()
        self._study_filter.addItem("All Studies", None)
        for sid, short_id in self._studies:
            self._study_filter.addItem(short_id, sid)
        idx = self._study_filter.findData(current)
        if idx >= 0:
            self._study_filter.setCurrentIndex(idx)
        self._study_filter.blockSignals(False)

    def refresh(self):
        self._load_studies()

        filters = {}
        study_id = self._study_filter.currentData()
        if study_id:
            filters["study_id"] = study_id
        pid_text = self._pid_search.text().strip()
        if pid_text:
            filters["pid_like"] = pid_text

        with get_session() as session:
            service = ParticipantService(session)
            participants, total = service.search(
                filters, page=self._page, page_size=PAGE_SIZE
            )
            data = [
                (
                    p.pid,
                    p.study_id,
                    str(p.age) if p.age else "",
                    p.gender or "",
                    p.population or "",
                    p.disease or "",
                    p.site_name or "",
                    p.cohort_name or "",
                    ", ".join(sorted(set(
                        s.visit_code for s in p.samples if s.visit_code
                    ))),
                    p.notes or "",
                    str(p.created_at.date()) if p.created_at else "",
                    p.id,
                )
                for p in participants
            ]

        study_map = {sid: short for sid, short in self._studies}

        self._table.setSortingEnabled(False)
        self._table.setRowCount(len(data))
        for row_idx, row_data in enumerate(data):
            display = list(row_data[:-1])
            display[1] = study_map.get(row_data[1], str(row_data[1]))

            for col_idx, value in enumerate(display):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter)
                self._table.setItem(row_idx, col_idx, item)

            self._table.item(row_idx, 0).setData(Qt.ItemDataRole.UserRole, row_data[-1])

        self._table.setSortingEnabled(True)
        self._btn_export.setEnabled(total > 0)
        self._lbl_count.setText(
            f"{total} participant{'s' if total != 1 else ''}  "
            f"(showing {len(data)})"
        )
        self._pagination.set_page(self._page, total, PAGE_SIZE)

    @slot_safe
    def _on_export(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Participants", "", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        filters = {}
        study_id = self._study_filter.currentData()
        if study_id:
            filters["study_id"] = study_id
        pid_text = self._pid_search.text().strip()
        if pid_text:
            filters["pid_like"] = pid_text

        with get_session() as session:
            service = ParticipantService(session)
            participants, _ = service.search(filters, page=1, page_size=9999)
            study_map = {sid: short for sid, short in self._studies}
            rows = [
                (
                    p.pid,
                    study_map.get(p.study_id, str(p.study_id)),
                    p.age or "",
                    p.gender or "",
                    p.population or "",
                    p.disease or "",
                    p.site_name or "",
                    p.cohort_name or "",
                    ", ".join(sorted(set(s.visit_code for s in p.samples if s.visit_code))),
                    p.notes or "",
                    str(p.created_at.date()) if p.created_at else "",
                )
                for p in participants
            ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Participants"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col_idx, col_name in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        wb.save(file_path)
        QMessageBox.information(
            self, "Export complete",
            f"Exported {len(rows)} participant(s) to:\n{file_path}"
        )

    @slot_safe
    def _on_new(self):
        from app.ui.dialogs.participant_dialog import ParticipantDialog
        study_id = self._study_filter.currentData()
        dlg = ParticipantDialog(self, preselect_study_id=study_id)
        if dlg.exec():
            self.refresh()

    @slot_safe
    def _on_import(self):
        from app.ui.dialogs.excel_import_dialog import ExcelImportDialog
        dlg = ExcelImportDialog(self)
        if dlg.exec():
            self.refresh()

    @slot_safe
    def _on_edit(self):
        row = self._table.currentRow()
        if row < 0:
            return
        participant_id = self._table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        from app.ui.dialogs.participant_dialog import ParticipantDialog
        dlg = ParticipantDialog(self, participant_id=participant_id)
        if dlg.exec():
            self.refresh()
