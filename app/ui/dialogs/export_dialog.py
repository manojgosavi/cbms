"""
Export to Excel dialog — exports search results with full Excel column structure.

Excel columns (in order):
PID, Age, Gender, Population, Disease, Visit Code, Visit Time, Date Collected,
Site Name, Visit Name, Sample Type, Cohort Name, Aliquot ID,
Freezer / Tank, Container, Slot Position, Shelf, Rack, Position,
Discrepancy Remark, Discrepancy For
"""

from __future__ import annotations

from typing import List, Optional

from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (
    QDialog, QDialogButtonBox, QFileDialog,
    QLabel, QVBoxLayout, QCheckBox, QHBoxLayout
)

from app.core.services.search_service import SearchResult
from app.utils.exception_handler import slot_safe


class ExportDialog(QDialog):

    def __init__(self, parent=None, results: List[SearchResult] = None):
        super().__init__(parent)
        self._results = results or []
        self.setWindowTitle("Export Search Results to Excel")
        self.setMinimumWidth(340)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        layout.addWidget(QLabel("Select columns to include:"))
        layout.addSpacing(8)

        # Checkboxes for column groups
        self._inc_participant = QCheckBox("Participant info (PID, Age, Gender, Population, Disease)")
        self._inc_participant.setChecked(True)
        self._inc_sample       = QCheckBox("Sample & aliquot details (Visit Code/Time, Date, Sample Type, Aliquot ID)")
        self._inc_sample.setChecked(True)
        self._inc_location     = QCheckBox("Storage location (Freezer, Container, Rack, Drawer, Box, Position)")
        self._inc_location.setChecked(True)
        self._inc_discrepancy  = QCheckBox("Discrepancy / QC notes")
        self._inc_discrepancy.setChecked(True)

        layout.addWidget(self._inc_participant)
        layout.addWidget(self._inc_sample)
        layout.addWidget(self._inc_location)
        layout.addWidget(self._inc_discrepancy)
        layout.addSpacing(12)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Save).setText("Choose file & export…")
        btns.accepted.connect(self._on_export)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    @slot_safe
    def _on_export(self):
        """Export to Excel file with exact Excel schema."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Error", "openpyxl not installed. Install with: pip install openpyxl")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel file", "", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        # Build header row based on selections
        headers = ["PID", "Age", "Gender", "Population", "Disease", "Visit Code", "Visit Time",
                   "Date Collected", "Site Name", "Visit Name", "Sample Type", "Cohort Name", "Aliquot ID"]

        if self._inc_location.isChecked():
            headers += ["Freezer / Tank", "Container", "Slot Position", "Shelf", "Rack", "Position"]

        if self._inc_discrepancy.isChecked():
            headers += ["Discrepancy Remark", "Discrepancy For"]

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"

        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data rows
        for row_idx, r in enumerate(self._results, 2):
            # Position label (grid position from row/col)
            pos_label = ""
            if r.position_row is not None and r.position_col is not None:
                col_label = chr(ord('A') + r.position_col) if r.position_col < 26 else str(r.position_col)
                pos_label = f"{r.position_row + 1}{col_label}"

            # Standard columns (always included)
            row_data = [
                r.pid,
                r.age,
                r.sex,  # Gender
                r.cohort,  # Population (note: search result has 'cohort' field)
                r.disease,
                "",  # Visit Code (not in search result, would need separate query)
                "",  # Visit Time
                str(r.collection_date) if r.collection_date else "",
                r.site_name,
                "",  # Visit Name (not in search result)
                r.sample_type,
                r.site_name,  # Cohort Name (placeholder, needs actual cohort)
                r.aliquot_id,
            ]

            if self._inc_location.isChecked():
                # Storage hierarchy: Freezer/Tank, Container, Slot Position, Shelf, Rack, Position
                row_data += [
                    r.freezer_name,                    # Freezer / Tank
                    r.compartment_name or r.rack_name or r.drawer_name,  # Container
                    r.box_name,                        # Slot Position
                    r.compartment_name or "",          # Shelf (mapped to compartment)
                    r.rack_name or "",                 # Rack
                    pos_label,                         # Position
                ]

            if self._inc_discrepancy.isChecked():
                row_data += [
                    r.discrepancy_remark or "",
                    "",  # Discrepancy For (would need separate field)
                ]

            # Write row
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        wb.save(file_path)

        from PyQt6.QtWidgets import QMessageBox
        QMessageBox.information(self, "Success", f"Exported {len(self._results)} records to {file_path}")
        self.accept()
