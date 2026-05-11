"""
Phase 5 tests — admin service and catalogue.
"""

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.models.models import Base
from app.core.services.auth_service import seed_admin, login, app_session, register_user
from app.core.services.admin_service import AdminService
from app.core.services.catalogue_service import CatalogueService
from app.core.services.study_service import StudyService
from app.core.services.participant_service import ParticipantService
from app.core.services.sample_service import SampleService
from app.config import Role


@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    seed_admin(session)
    session.commit()
    login(session, "admin", "Admin@1234")
    yield session
    session.close()
    app_session.logout()


@pytest.fixture
def full_data(db):
    study_svc = StudyService(db)
    _, _, study = study_svc.create_study("CAT", "Catalogue Study")
    db.commit()
    p_svc = ParticipantService(db)
    _, _, p1 = p_svc.register_participant(study.id, "AB", age=30,
                                           sex="Male", disease="TB")
    _, _, p2 = p_svc.register_participant(study.id, "CD", age=45,
                                           sex="Female", disease="CVD")
    db.commit()
    s_svc = SampleService(db)
    s_svc.register_sample(p1.id, study.id, "Serum",      num_aliquots=3)
    s_svc.register_sample(p1.id, study.id, "ED Plasma",  num_aliquots=2)
    s_svc.register_sample(p2.id, study.id, "Serum",      num_aliquots=1)
    s_svc.register_sample(p2.id, study.id, "EDTA PBMC",  num_aliquots=4)
    db.commit()
    return study, p1, p2


# ══════════════════════════════════════════════════════════════════════════
# ADMIN SERVICE
# ══════════════════════════════════════════════════════════════════════════

class TestAdminService:

    def test_get_all_users(self, db):
        svc = AdminService(db)
        users = svc.get_all_users()
        assert len(users) >= 1
        assert any(u.username == "admin" for u in users)

    def test_approve_user(self, db):
        register_user(db, "newuser", "new@lab.org", "pass", Role.LAB_TECH,
                      auto_approve=False)
        db.commit()
        from app.core.models.models import User
        user = db.query(User).filter_by(username="newuser").first()
        svc = AdminService(db)
        ok, msg = svc.approve_user(user.id)
        assert ok, msg
        db.refresh(user)
        assert user.is_approved

    def test_cannot_approve_already_approved(self, db):
        from app.core.models.models import User
        admin = db.query(User).filter_by(username="admin").first()
        ok, msg = AdminService(db).approve_user(admin.id)
        assert not ok

    def test_update_user_role(self, db):
        register_user(db, "techuser", "tech@lab.org", "pass",
                      Role.LAB_TECH, auto_approve=True)
        db.commit()
        from app.core.models.models import User
        user = db.query(User).filter_by(username="techuser").first()
        ok, msg = AdminService(db).update_user(user.id, role=Role.MANAGER)
        assert ok, msg
        db.refresh(user)
        assert user.role == Role.MANAGER

    def test_cannot_change_own_role(self, db):
        from app.core.models.models import User
        admin = db.query(User).filter_by(username="admin").first()
        ok, msg = AdminService(db).update_user(admin.id, role=Role.LAB_TECH)
        assert not ok
        assert "own role" in msg.lower()

    def test_reset_password(self, db):
        from app.core.models.models import User
        admin = db.query(User).filter_by(username="admin").first()
        ok, msg = AdminService(db).reset_password(admin.id, "NewPass@123")
        assert ok, msg

    def test_reset_password_too_short(self, db):
        from app.core.models.models import User
        admin = db.query(User).filter_by(username="admin").first()
        ok, msg = AdminService(db).reset_password(admin.id, "short")
        assert not ok

    def test_delete_user(self, db):
        register_user(db, "todelete", "del@lab.org", "pass",
                      Role.LAB_TECH, auto_approve=True)
        db.commit()
        from app.core.models.models import User
        user = db.query(User).filter_by(username="todelete").first()
        ok, msg = AdminService(db).delete_user(user.id)
        assert ok, msg
        assert db.query(User).filter_by(username="todelete").first() is None

    def test_cannot_delete_self(self, db):
        from app.core.models.models import User
        admin = db.query(User).filter_by(username="admin").first()
        ok, msg = AdminService(db).delete_user(admin.id)
        assert not ok

    def test_add_custom_field(self, db):
        svc = AdminService(db)
        ok, msg = svc.add_custom_field("hiv_status", "HIV Status", "select")
        assert ok, msg

    def test_duplicate_custom_field_rejected(self, db):
        svc = AdminService(db)
        svc.add_custom_field("dup_field", "Dup", "text")
        ok, msg = svc.add_custom_field("dup_field", "Dup2", "text")
        assert not ok

    def test_audit_log_written_on_approve(self, db):
        register_user(db, "audituser", "a@lab.org", "pass",
                      Role.LAB_TECH, auto_approve=False)
        db.commit()
        from app.core.models.models import User, AuditLog
        user = db.query(User).filter_by(username="audituser").first()
        AdminService(db).approve_user(user.id)
        db.commit()
        log = db.query(AuditLog).filter_by(entity_type="User",
                                            entity_id=str(user.id)).first()
        assert log is not None


# ══════════════════════════════════════════════════════════════════════════
# CATALOGUE SERVICE
# ══════════════════════════════════════════════════════════════════════════

class TestCatalogueService:

    def test_generate_catalogue(self, db, full_data):
        study, p1, p2 = full_data
        svc = CatalogueService(db)
        rows, col_headers = svc.generate(study_id=study.id)
        assert len(rows) == 2
        assert set(col_headers) == {"Serum", "ED Plasma", "EDTA PBMC"}

    def test_pivot_counts_correct(self, db, full_data):
        study, p1, p2 = full_data
        svc = CatalogueService(db)
        rows, headers = svc.generate(study_id=study.id)
        row_p1 = next(r for r in rows if r.pid == p1.pid)
        assert row_p1.sample_counts["Serum"]     == 3
        assert row_p1.sample_counts["ED Plasma"] == 2
        assert row_p1.total_aliquots             == 5

    def test_total_aliquots_correct(self, db, full_data):
        study, p1, p2 = full_data
        svc = CatalogueService(db)
        rows, _ = svc.generate(study_id=study.id)
        grand_total = sum(r.total_aliquots for r in rows)
        assert grand_total == 10   # 3+2+1+4

    def test_all_studies_no_filter(self, db, full_data):
        svc = CatalogueService(db)
        rows, _ = svc.generate()
        assert len(rows) >= 2

    def test_export_to_excel(self, db, full_data, tmp_path):
        study, _, _ = full_data
        svc = CatalogueService(db)
        rows, headers = svc.generate(study_id=study.id)
        path = str(tmp_path / "test_catalogue.xlsx")
        svc.export_to_excel(rows, headers, path)
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Header row + data rows + totals row
        assert ws.max_row == len(rows) + 2
        # First header should be PID
        assert ws.cell(1, 1).value == "PID"
